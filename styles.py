# -*- coding: utf-8 -*-
import time
import win32gui
import win32con
from openpyxl import Workbook, styles
import xlwings as xw
import settings as st


def add_styles(dft: tuple):
    """
    dft = namedtuple('final_assembly', ['e_rdc', 'e_sjc', 'i_rdc', 'i_sjc'])
    """
    fields = list(dft._fields)
    wb = Workbook()

    title_style = styles.NamedStyle(name='title_style')
    title_style.font = styles.Font(
        name='微软雅黑',
        size=11,
        bold=True
    )
    title_style.alignment = styles.Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
    )
    title_style.number_format = '@'
    wb.add_named_style(title_style)

    for enum, df in enumerate(dft):
        if enum == 0:
            ws = wb.active
            ws.title = fields[enum]
        else:
            ws = wb.create_sheet(title=fields[enum])
        multi_column = df.columns
        # ------------------------------------
        columns = ['序号']
        d_type = ['int']
        columns.extend(multi_column.get_level_values(0))
        d_type.extend(multi_column.get_level_values(3))
        columns_size = multi_column.size + 1
        type_set = {'int': '0', 'str': '@', 'float': '0.00', '%': '0%'}
        a_upper = 65
        freeze_panes = [None, True]
        for ii in range(columns_size):
            col = chr(a_upper + ii % 26)
            if ii // 26 > 0:
                col = chr(a_upper + ii // 26 - 1) + col
            for _ in st.FEATURE_PROPERTY:
                jj = st.FEATURE_PROPERTY[_]
                if jj['name'] == columns[ii]:
                    w = jj.get('width', 6)
                    wrap = jj.get('wrap_text', False)
                    alm = jj.get('alignment', 'center')
                    bold = jj.get('bold', False)
                    nf = type_set[d_type[ii]]
                    ws.column_dimensions[col].width = w
                    ws.column_dimensions[col].font = styles.Font(
                        name='微软雅黑',
                        size=10,
                        bold=bold
                    )
                    ws.column_dimensions[col].alignment = styles.Alignment(
                        horizontal=alm,
                        vertical='center',
                        wrap_text=wrap
                    )
                    ws.column_dimensions[col].number_format = nf
                    # 标记需要冻结的列, 只需标注第一个出现的即可
                    if freeze_panes[1] & jj.get('freeze_panes', False):
                        freeze_panes[0] = col + str(2)
                        freeze_panes[1] = False
            # 没有在列属性中设置冻结, 显式冻结首行
            freeze_panes[0] = 'A2'

        # 不能对row_dimensions设置style
        for ii in range(columns_size):
            ws.cell(row=1, column=ii+1).style = 'title_style'
        columns_size -= 1
        cell = chr(a_upper + columns_size % 26)
        if columns_size // 26 > 0:
            cell = chr(a_upper + columns_size // 26 - 1) + cell
        ws.auto_filter.ref = 'A1:' + cell + str(df.index.size+1)
        ws.freeze_panes = freeze_panes[0]
        # column = 'A:B,E:H'
        # ws.range(column).column_width = width[ii]
    # ------------------------------------
    file_path = st.FILE_GENERATED_PATH
    wb.save(file_path)
    wb.close()
    # print(time.time() - old_time)
    # -----------------------------------
    add_data(dft, file_path)
    # -----------------------------------


def add_data(dft, file_path):
    """
    打开openpyxl创建的Excel表格, 用xlwings填充数据, 看起来比较直观
    """
    visible = st.SHOW_DOC_AFTER_GENERATED
    app = xw.App(visible=visible, add_book=False)
    wb = app.books.open(file_path)
    for enum, df in enumerate(dft):
        ws = wb.sheets[enum]
        print('********************', ws.name)
        df = dft[enum]
        df = df.droplevel(level=[1, 2, 3], axis=1)
        ws.range('A1').value = df
    wb.save()
    if visible:
        # 通过xlwings app获取窗口句柄, 再使用win32接口最大化 最小化是: SW_SHOWMINIMIZED
        hwnd = app.hwnd
        win32gui.ShowWindow(hwnd, win32con.SW_SHOWMAXIMIZED)
    else:
        wb.close()
        app.quit()
    print('------fulfill the task!------')
